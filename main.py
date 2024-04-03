import openpyxl

import requests
import json
import datetime


with open('login_data.txt', 'r', encoding='utf-8') as file:
    data = [i for i in file]
client_id = data[0].replace('\n', '')
api_key = data[1].replace('\n', '')
headers = {
    "Client-Id": client_id,
    "Api-Key": api_key
}

url = 'https://api-seller.ozon.ru/v3/posting/fbs/list'
yesterday = datetime.date.today()- datetime.timedelta(days=1)
after_10_days = yesterday + datetime.timedelta(days=10)
payload = json.dumps({
    "dir": "ASC",
    "filter": {

        "order_id": 0,
        "since": f"{yesterday}T11:47:39.878Z",
        "to": f"{after_10_days}T11:47:39.878Z",
    },
    "limit": 1000,


})
resp_data = requests.post(url, headers=headers, data=payload).json()

today = datetime.date.today()
tomorrow = today + datetime.timedelta(days=1)

wb = openpyxl.Workbook('Результат.xlsx')
wb.save('Результат.xlsx')
book = openpyxl.load_workbook('Результат.xlsx')
sheet=book.active
row=1
for i in resp_data['result']['postings']:
    date_time_str=i['shipment_date'][:10]
    date= datetime.datetime.strptime(date_time_str, '%Y-%m-%d')
    date=date.date()
    if date==tomorrow:
        sku=i['products'][0]['offer_id']
        quantity=i['products'][0]['quantity']
        sheet.cell(column=1, row=row).value = sku
        sheet.cell(column=2, row=row).value = quantity
        sheet.cell(column=3, row=row).value = date
        row+=1
        book.save('Результат.xlsx')
book.close()


